import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows  # Correct import
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta

# Initialize Workbook
wb = Workbook()

# Remove default sheet
default_sheet = wb.active
wb.remove(default_sheet)

# Function to auto-adjust column widths
def auto_adjust_column_widths(sheet):
    for column_cells in sheet.columns:
        max_length = 0
        column = column_cells[0].column_letter
        for cell in column_cells:
            try:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass
        adjusted_width = (max_length + 2) if max_length < 50 else 50
        sheet.column_dimensions[column].width = adjusted_width

# Define Experiment Parameters
start_date = datetime(2024, 1, 1)
num_weeks = 6
groups = {
    "Control": {"BSF": 0, "Soy": 33, "Carbs": 67},
    "Low BSF": {"BSF": 10, "Soy": 23, "Carbs": 67},
    "Moderate BSF": {"BSF": 15, "Soy": 18, "Carbs": 67},
    "High BSF": {"BSF": 20, "Soy": 13, "Carbs": 67}
}
num_chickens_per_group = 25

# Generate Bird IDs
bird_ids = {}
for group in groups.keys():
    bird_ids[group] = [f"{group[0]}{i+1}" for i in range(num_chickens_per_group)]

# 1. Feed Consumption Sheet
feed_consumption = []
for week in range(num_weeks):
    current_date = (start_date + timedelta(weeks=week)).strftime("%Y-%m-%d")
    for group, composition in groups.items():
        # Assume each group is allocated 3 kg per cycle, distributed weekly
        weekly_feed_total = 3 / num_weeks  # kg per group per week
        feed_offered = weekly_feed_total * 1000  # grams
        # Assume feed refusal rate increases with BSF inclusion
        refusal_rate = 30 + (composition["BSF"] * 1.5)
        feed_refused = refusal_rate + np.random.normal(0, 5)
        feed_refused = max(0, feed_refused)  # Ensure non-negative
        notes = ""
        if composition["BSF"] > 15:
            if np.random.rand() < 0.2:
                notes = "Slight refusal observed"
        elif composition["BSF"] > 10:
            if np.random.rand() < 0.1:
                notes = "Minor leftovers"
        else:
            if np.random.rand() < 0.05:
                notes = "Good consumption"
        feed_consumption.append({
            "Date": current_date,
            "Group": group,
            "Feed Offered (g)": round(feed_offered, 2),
            "Feed Refused (g)": round(feed_refused, 2),
            "Notes": notes
        })

feed_df = pd.DataFrame(feed_consumption)

feed_sheet = wb.create_sheet("Feed Consumption")
for r in dataframe_to_rows(feed_df, index=False, header=True):
    feed_sheet.append(r)
auto_adjust_column_widths(feed_sheet)

# 2. Feed Composition Breakdown Sheet
# Here we break down the consumed feed into BSF, Soy, and Carbs per group and date
# Consumed = Offered - Refused
feed_breakdown = []
for _, row in feed_df.iterrows():
    consumed = row["Feed Offered (g)"] - row["Feed Refused (g)"]
    group_name = row["Group"]
    composition = groups[group_name]
    # Calculate grams of each ingredient consumed based on composition percentages
    total_protein = composition["BSF"] + composition["Soy"]  # total protein portion = 33
    # total_protein is always 33, but let's keep it generic in case of future changes
    total_percentage = composition["BSF"] + composition["Soy"] + composition["Carbs"]  # should be 100
    # Determine grams of each ingredient
    # Note: composition values are percentages
    bsf_g = (composition["BSF"] / 100.0) * consumed
    soy_g = (composition["Soy"] / 100.0) * consumed
    carbs_g = (composition["Carbs"] / 100.0) * consumed

    feed_breakdown.append({
        "Date": row["Date"],
        "Group": group_name,
        "BSF Consumed (g)": round(bsf_g, 2),
        "Soy Consumed (g)": round(soy_g, 2),
        "Carbs Consumed (g)": round(carbs_g, 2)
    })

breakdown_df = pd.DataFrame(feed_breakdown)
composition_sheet = wb.create_sheet("Feed Composition Breakdown")
for r in dataframe_to_rows(breakdown_df, index=False, header=True):
    composition_sheet.append(r)
auto_adjust_column_widths(composition_sheet)

# 3. Weight Measurements Sheet
weight_measurements = []
initial_weights = {group: 800 + np.random.normal(0, 20) for group in groups.keys()}
current_weights = {group: {bird: initial_weights[group] for bird in bird_ids[group]} for group in groups.keys()}

for week in range(num_weeks):
    current_date = (start_date + timedelta(weeks=week)).strftime("%Y-%m-%d")
    for group, birds in bird_ids.items():
        for bird in birds:
            base_gain = 50
            bsfl_factor = (groups[group]["BSF"] / 10)
            weight_gain = base_gain + (bsfl_factor * 10) + np.random.normal(0, 5)
            weight_gain = max(0, weight_gain)
            new_weight = current_weights[group][bird] + weight_gain
            current_weights[group][bird] = new_weight
            notes = ""
            if groups[group]["BSF"] > 15:
                if np.random.rand() < 0.1:
                    notes = "Feather ruffling observed"
            elif groups[group]["BSF"] > 10:
                if np.random.rand() < 0.05:
                    notes = "Active and healthy"
            else:
                if np.random.rand() < 0.02:
                    notes = "Good growth"
            weight_measurements.append({
                "Date": current_date,
                "Group": group,
                "Bird ID": bird,
                "Weight (g)": round(new_weight, 2),
                "Notes": notes
            })

weight_df = pd.DataFrame(weight_measurements)
weight_sheet = wb.create_sheet("Weight Measurements")
for r in dataframe_to_rows(weight_df, index=False, header=True):
    weight_sheet.append(r)
auto_adjust_column_widths(weight_sheet)

# 4. Health Observations Sheet
health_observations = []
for week in range(num_weeks):
    current_date = (start_date + timedelta(weeks=week)).strftime("%Y-%m-%d")
    for group, birds in bird_ids.items():
        for bird in birds:
            if np.random.rand() < 0.05:
                observation_type = np.random.choice(["Illness", "Mortality", "Behavior"])
                if observation_type == "Illness":
                    description = "Lethargic behavior observed"
                    action_taken = "Isolated and treated with electrolytes"
                    outcome = "Improving"
                elif observation_type == "Mortality":
                    description = "Unexpected death"
                    action_taken = "Investigated potential causes"
                    outcome = "Unknown"
                else:
                    description = "Increased pecking at feathers"
                    action_taken = "Separated affected bird temporarily"
                    outcome = "Behavior normalized"
                health_observations.append({
                    "Date": current_date,
                    "Group": group,
                    "Bird ID": bird,
                    "Observation Type": observation_type,
                    "Description": description,
                    "Action Taken": action_taken,
                    "Outcome": outcome
                })

health_df = pd.DataFrame(health_observations)
health_sheet = wb.create_sheet("Health Observations")
for r in dataframe_to_rows(health_df, index=False, header=True):
    health_sheet.append(r)
auto_adjust_column_widths(health_sheet)

# 5. Cost and Inventory Sheet
cost_inventory = []
cost_inventory.append({
    "Item/Ingredient": "BSF Powder",
    "Type": "BSFL",
    "Unit Cost (KWD/ton)": 120,
    "Quantity (kg)": 200,
    "Date of Purchase": "2024-01-01",
    "Supplier": "InsectFarm Co.",
    "Notes": "High quality, organic"
})
cost_inventory.append({
    "Item/Ingredient": "Soy Meal",
    "Type": "Soy",
    "Unit Cost (KWD/ton)": 230,
    "Quantity (kg)": 300,
    "Date of Purchase": "2024-01-01",
    "Supplier": "AgroSupplies",
    "Notes": "Standard grade"
})
cost_inventory.append({
    "Item/Ingredient": "Corn Mix",
    "Type": "Carbs",
    "Unit Cost (KWD/ton)": 150,
    "Quantity (kg)": 500,
    "Date of Purchase": "2024-01-01",
    "Supplier": "GrainCorp",
    "Notes": "Bulk purchase"
})

for week in range(num_weeks):
    purchase_date = (start_date + timedelta(weeks=week)).strftime("%Y-%m-%d")
    if week % 3 == 0 and week != 0:
        cost_inventory.append({
            "Item/Ingredient": "BSF Powder",
            "Type": "BSFL",
            "Unit Cost (KWD/ton)": 120,
            "Quantity (kg)": 100,
            "Date of Purchase": purchase_date,
            "Supplier": "InsectFarm Co.",
            "Notes": "Restock for high BSF groups"
        })
    if week % 2 == 0:
        cost_inventory.append({
            "Item/Ingredient": "Soy Meal",
            "Type": "Soy",
            "Unit Cost (KWD/ton)": 230,
            "Quantity (kg)": 100,
            "Date of Purchase": purchase_date,
            "Supplier": "AgroSupplies",
            "Notes": "Restock for all groups"
        })
    if week % 1 == 0:
        cost_inventory.append({
            "Item/Ingredient": "Corn Mix",
            "Type": "Carbs",
            "Unit Cost (KWD/ton)": 150,
            "Quantity (kg)": 200,
            "Date of Purchase": purchase_date,
            "Supplier": "GrainCorp",
            "Notes": "Weekly restock"
        })

cost_df = pd.DataFrame(cost_inventory)
cost_sheet = wb.create_sheet("Cost and Inventory")
for r in dataframe_to_rows(cost_df, index=False, header=True):
    cost_sheet.append(r)
auto_adjust_column_widths(cost_sheet)

# 6. Protocol and Adjustments Sheet
protocol_adjustments = [
    {
        "Parameter": "Experiment Duration",
        "Details": "45 days total (6 weeks)",
        "Proposed Change": "",
        "Approval Status": "Approved",
        "Notes": ""
    },
    {
        "Parameter": "Feed Ratios",
        "Details": "Control: 0% BSF; Test Groups: 10%,15%,20% BSF",
        "Proposed Change": "",
        "Approval Status": "Approved",
        "Notes": ""
    },
    {
        "Parameter": "Weighing Frequency",
        "Details": "Weekly weigh-ins",
        "Proposed Change": "",
        "Approval Status": "Approved",
        "Notes": ""
    },
    {
        "Parameter": "Health Monitoring",
        "Details": "Weekly health checks with daily brief inspections",
        "Proposed Change": "",
        "Approval Status": "Approved",
        "Notes": ""
    }
]

protocol_df = pd.DataFrame(protocol_adjustments)
protocol_sheet = wb.create_sheet("Protocol and Adjustments")
for r in dataframe_to_rows(protocol_df, index=False, header=True):
    protocol_sheet.append(r)
auto_adjust_column_widths(protocol_sheet)

# Save the Workbook
wb.save("BSFL_Expdiidididieriment_Full_Dummy_Data.xlsx")
print("Excel workbook 'BSFL_Experddddiment_Full_Dummy_Data.xlsx' created successfully.")
