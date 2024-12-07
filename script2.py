from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# Create a new workbook and remove the default sheet
wb = Workbook()
if "Sheet" in wb.sheetnames:
    wb.remove(wb["Sheet"])

def style_headers(sheet, header_cells):
    """Apply bold, centered formatting to header cells and freeze top rows."""
    for cell in header_cells:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        # Light gray fill for headers for clarity
        cell.fill = PatternFill(start_color="DDDDDD", fill_type="solid")
    # Freeze the top row
    sheet.freeze_panes = sheet['A3']  # Freeze first two rows to keep instructions visible

def auto_adjust_column_widths(sheet):
    """Auto-adjust column widths based on cell contents."""
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            val = str(cell.value) if cell.value is not None else ""
            if len(val) > max_length:
                max_length = len(val)
        adjusted_width = max_length + 4
        sheet.column_dimensions[column].width = adjusted_width

def add_instructions_row(sheet, instructions_list, row_number=2):
    """Add a row of instructions below the headers to clarify data entry."""
    for col_num, instr in enumerate(instructions_list, start=1):
        cell = sheet.cell(row=row_number, column=col_num, value=instr)
        cell.font = Font(italic=True, color="555555")
        cell.alignment = Alignment(wrap_text=True, vertical='top')

# 0. Instructions Sheet
instructions_sheet = wb.create_sheet("Instructions")
instructions_sheet["A1"] = "How to Use This Workbook"
instructions_sheet["A1"].font = Font(bold=True, size=14)
instructions_content = [
    "This workbook is designed to help you track and analyze the BSFL poultry feed experiment.",
    "",
    "Sheets and Their Purposes:",
    "- Feed Consumption: Record daily feed offered, any refused, and calculate total consumed.",
    "- Weight Measurements: Track weekly weights of sampled birds and final weights.",
    "- Health Observations: Note any illnesses, mortalities, or unusual behavior and the actions taken.",
    "- Cost and Inventory: Keep track of ingredient costs, amounts purchased, and supplier details.",
    "- Protocol and Adjustments: Outline the experiment protocol, note any approved changes, and record important parameters.",
    "",
    "General Tips:",
    "- Fill data below the example rows provided.",
    "- Update daily and weekly as instructed.",
    "- For urgent issues, record details in the appropriate sheet and report immediately as per your instructions.",
    "",
    "Feel free to add more sheets or columns if needed, but keep the structure for consistent data analysis."
]

for i, line in enumerate(instructions_content, start=2):
    instructions_sheet.cell(row=i, column=1, value=line)

auto_adjust_column_widths(instructions_sheet)

# 1. Feed Consumption Tracking Sheet
feed_sheet = wb.create_sheet("Feed Consumption")
feed_headers = [
    "Date (YYYY-MM-DD)", 
    "Group (e.g., Control, 25% BSFL, 35% BSFL)", 
    "Type of Feed (BSFL/Soy/Mix)", 
    "Amount Offered (kg)",
    "Amount Refused (kg)", 
    "Total Consumed (kg) [=Offered - Refused]", 
    "Notes (e.g., weather, irregularities)"
]

feed_instructions = [
    "Enter the date of feeding.",
    "Enter the group name or ID.",
    "Specify the feed type or ratio.",
    "Amount of feed offered to the group.",
    "Amount of feed not consumed by the next feeding time.",
    "Calculate or verify: Offered - Refused.",
    "Any additional remarks."
]

for col_num, header in enumerate(feed_headers, start=1):
    feed_sheet.cell(row=1, column=col_num, value=header)

add_instructions_row(feed_sheet, feed_instructions)
style_headers(feed_sheet, feed_sheet[1])

# Add example entry row (row 3)
example_feed_data = [
    "2024-12-10", "25% BSFL", "25% BSFL + rest soy",
    5.0, 0.5, "=D3-E3", "Slightly cooler day, birds active"
]
for col_num, value in enumerate(example_feed_data, start=1):
    feed_sheet.cell(row=3, column=col_num, value=value)

auto_adjust_column_widths(feed_sheet)

# 2. Weight Measurements Sheet
weight_sheet = wb.create_sheet("Weight Measurements")
weight_headers = [
    "Date (YYYY-MM-DD)", "Group", "Bird ID", "Weight (g)",
    "Average Weight (Group) [Calculate after all sample birds weighed]", 
    "Cumulative Gain (g) [final - initial]", "Notes"
]

weight_instructions = [
    "Enter the date of weighing.",
    "Enter the group name or ID.",
    "Unique ID for each bird sampled.",
    "Record the bird's weight in grams.",
    "After recording all birds for the day, calculate the average weight for the group.",
    "Track weight gain from previous measurement or from start.",
    "Any observations (e.g., bird looks healthy, slight limping, etc.)."
]

for col_num, header in enumerate(weight_headers, start=1):
    weight_sheet.cell(row=1, column=col_num, value=header)

add_instructions_row(weight_sheet, weight_instructions)
style_headers(weight_sheet, weight_sheet[1])

# Example entry row
example_weight_data = [
    "2024-12-17", "25% BSFL", "Bird_01",
    820, "", "", "Bird active and feeding well"
]
for col_num, value in enumerate(example_weight_data, start=1):
    weight_sheet.cell(row=3, column=col_num, value=value)

auto_adjust_column_widths(weight_sheet)

# 3. Health Observations Sheet
health_sheet = wb.create_sheet("Health Observations")
health_headers = [
    "Date (YYYY-MM-DD)", "Group", "Bird ID (if applicable)", "Observation Type (Illness/Mortality/Behavior)",
    "Description of Issue", "Action Taken", "Outcome", "Notes"
]

health_instructions = [
    "Date of observation.",
    "Group affected.",
    "Bird ID if it's an individual bird; leave blank if group-level note.",
    "Type of issue: Illness, Mortality, or Behavior change.",
    "Describe the symptom or issue in detail.",
    "What action did you take (e.g., isolation, veterinary check)?",
    "Outcome or follow-up required.",
    "Any extra notes."
]

for col_num, header in enumerate(health_headers, start=1):
    health_sheet.cell(row=1, column=col_num, value=header)

add_instructions_row(health_sheet, health_instructions)
style_headers(health_sheet, health_sheet[1])

# Example entry
example_health_data = [
    "2024-12-20", "35% BSFL", "Bird_05", "Illness",
    "Sneezing, lethargic", "Isolated, provided electrolytes", "Monitoring required", "Vet visit scheduled tomorrow"
]
for col_num, value in enumerate(example_health_data, start=1):
    health_sheet.cell(row=3, column=col_num, value=value)

auto_adjust_column_widths(health_sheet)

# 4. Cost and Inventory Sheet
cost_sheet = wb.create_sheet("Cost and Inventory")
cost_headers = [
    "Item/Ingredient", "Type (Soy/BSFL/Other)", "Unit Cost (KWD/ton)",
    "Quantity Purchased (kg)", "Total Cost (KWD) [= (Unit Cost/1000)*Quantity]", "Date of Purchase", "Supplier", "Notes"
]

cost_instructions = [
    "Name of the ingredient or item.",
    "Type/category of the ingredient.",
    "Cost per ton (1,000 kg).",
    "How many kg purchased.",
    "Calculate total cost: (Unit Cost/1000)*Quantity.",
    "When it was purchased.",
    "From whom you bought it.",
    "Any additional notes (quality, discounts, etc.)."
]

for col_num, header in enumerate(cost_headers, start=1):
    cost_sheet.cell(row=1, column=col_num, value=header)

add_instructions_row(cost_sheet, cost_instructions)
style_headers(cost_sheet, cost_sheet[1])

# Example entry
example_cost_data = [
    "BSFL Meal", "BSFL", 120, 200, "=C3/1000*D3", "2024-12-08", "InsectFarm Co.", "Good quality batch"
]
for col_num, value in enumerate(example_cost_data, start=1):
    cost_sheet.cell(row=3, column=col_num, value=value)

auto_adjust_column_widths(cost_sheet)

# 5. Protocol and Adjustments Sheet
protocol_sheet = wb.create_sheet("Protocol and Adjustments")
protocol_headers = [
    "Parameter/Aspect", "Details/Instructions", "Proposed Changes", "Approval Status", "Notes"
]

protocol_instructions = [
    "The parameter or aspect of the experiment (e.g., feeding schedule, ratio).",
    "Initial instructions or protocol details.",
    "Any proposed change to this parameter.",
    "Approved, Pending, or Rejected by the project lead.",
    "Additional notes or justification for changes."
]

for col_num, header in enumerate(protocol_headers, start=1):
    protocol_sheet.cell(row=1, column=col_num, value=header)

add_instructions_row(protocol_sheet, protocol_instructions)
style_headers(protocol_sheet, protocol_sheet[1])

# Pre-fill some standard protocol guidelines
example_protocol_data = [
    ["Experiment Duration", "45 days total from day-old chicks to slaughter weight.", "", "Approved", ""],
    ["Feed Ratios", "Control: 0% BSFL, Group 1: 25% BSFL, Group 2: 35% BSFL", "Increase Group 2 to 40% if performance is good", "Pending", "Will decide after first 2 weeks of data"],
    ["Weighing Schedule", "Weigh 10 birds/group weekly; final weigh all at Day 45", "", "Approved", ""]
]

start_row = 3
for row_data in example_protocol_data:
    for col_num, value in enumerate(row_data, start=1):
        protocol_sheet.cell(row=start_row, column=col_num, value=value)
    start_row += 1

auto_adjust_column_widths(protocol_sheet)

# Save the workbook
wb.save("BSFL_Experiment_Data_Tracking_filled.xlsx")
print("Excel workbook created successfully: BSFL_Experiment_Data_Tracking_filled.xlsx")

