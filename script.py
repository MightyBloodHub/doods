from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

# Create a new workbook and remove the default sheet
wb = Workbook()
if "Sheet" in wb.sheetnames:
    wb.remove(wb["Sheet"])
#jddjjd
# Define a function to style header cells
def style_headers(sheet, header_cells):
    for cell in header_cells:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
    # Freeze the top row
    sheet.freeze_panes = sheet['A2']

def auto_adjust_column_widths(sheet):
    # Auto-adjust column widths based on the length of the longest cell in each column
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            val = str(cell.value) if cell.value is not None else ""
            if len(val) > max_length:
                max_length = len(val)
        # Add a little extra space
        adjusted_width = max_length + 2
        sheet.column_dimensions[column].width = adjusted_width

# 1. Feed Consumption Tracking Sheet
feed_sheet = wb.create_sheet("Feed Consumption")
feed_headers = [
    "Date", "Group", "Type of Feed (BSFL/Soy/Other Mix)",
    "Amount Offered (kg)", "Amount Refused (kg)", "Total Consumed (kg)",
    "Notes"
]
for col_num, header in enumerate(feed_headers, start=1):
    feed_sheet.cell(row=1, column=col_num, value=header)

style_headers(feed_sheet, feed_sheet[1])
auto_adjust_column_widths(feed_sheet)

# 2. Weight Measurements Sheet
weight_sheet = wb.create_sheet("Weight Measurements")
weight_headers = [
    "Date", "Group", "Bird ID", "Weight (g)",
    "Average Weight (Group)", "Cumulative Gain (g)", "Notes"
]
for col_num, header in enumerate(weight_headers, start=1):
    weight_sheet.cell(row=1, column=col_num, value=header)

style_headers(weight_sheet, weight_sheet[1])
auto_adjust_column_widths(weight_sheet)

# 3. Health Observations Sheet
health_sheet = wb.create_sheet("Health Observations")
health_headers = [
    "Date", "Group", "Bird ID (if applicable)", "Observation Type (Illness/Mortality/Behavior)",
    "Description of Issue", "Action Taken", "Outcome", "Notes"
]
for col_num, header in enumerate(health_headers, start=1):
    health_sheet.cell(row=1, column=col_num, value=header)

style_headers(health_sheet, health_sheet[1])
auto_adjust_column_widths(health_sheet)

# 4. Cost and Inventory Sheet
cost_sheet = wb.create_sheet("Cost and Inventory")
cost_headers = [
    "Item/Ingredient", "Type (Soy/BSFL/Other)", "Unit Cost (KWD/ton)",
    "Quantity Purchased (kg)", "Total Cost (KWD)", "Date of Purchase", "Supplier", "Notes"
]
for col_num, header in enumerate(cost_headers, start=1):
    cost_sheet.cell(row=1, column=col_num, value=header)

style_headers(cost_sheet, cost_sheet[1])
auto_adjust_column_widths(cost_sheet)

# 5. Protocol and Adjustments Sheet
protocol_sheet = wb.create_sheet("Protocol and Adjustments")
protocol_headers = [
    "Parameter/Aspect", "Details/Instructions", "Proposed Changes", "Approval Status", "Notes"
]
for col_num, header in enumerate(protocol_headers, start=1):
    protocol_sheet.cell(row=1, column=col_num, value=header)

# Add some suggested rows in the Protocol sheet for clarity
protocol_sheet.cell(row=2, column=1, value="Experiment Duration")
protocol_sheet.cell(row=2, column=2, value="45 days total")
protocol_sheet.cell(row=3, column=1, value="Feed Ratios")
protocol_sheet.cell(row=3, column=2, value="Group 1: 25% BSFL replacement; Group 2: 35% BSFL replacement")
protocol_sheet.cell(row=4, column=1, value="Weighing Schedule")
protocol_sheet.cell(row=4, column=2, value="Weekly samples of 10 birds; final weigh all")

style_headers(protocol_sheet, protocol_sheet[1])
auto_adjust_column_widths(protocol_sheet)

# Save the workbook
wb.save("BSFL_Experiment_Data_Tracking.xlsx")

print("Excel workbook created successfully: BSFL_Experiment_Data_Tracking.xlsx")

